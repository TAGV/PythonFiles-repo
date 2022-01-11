import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill

#Classes in openpyxl
#Workbook--Worksheet--Cell


wb = openpyxl.load_workbook("/home/myubuntu/Desktop/Ecom_data.xlsx")
print(type(wb))     # <class 'openpyxl.workbook.workbook.Workbook'>

sheet_names = wb.sheetnames
print(sheet_names)

#print(wb.active.title)     #Get active sheet name

#Load the active sheet

sh1 = wb['Ecommerce_Data']

#Read the data
print(sh1['B2'].value)
print(sh1.cell(2,2).value)

#Looping and Printing
##Getting total no of rows and columns
row = sh1.max_row
column = sh1.max_column
print(row)
print(column)

#Printing the whole excel sheet
for i in range(1,row+1):
    for j in range(1,column+1):
        print(sh1.cell(i,j).value,end=' ')
    print()

#Writing to excel files
sh1.cell(1,1,value='Sr.No')

#Saving to a new file
wb.save("Report.xlsx")

print("============================================================================")

#Creating a excel sheet from scratch

cd = Workbook()

cd['Sheet'].title = "Automation Sheet"
print(cd.sheetnames)
print(cd.active.title)

#Create a new sheet in same workbook
cd.create_sheet('Logging')
print(cd.sheetnames)

print(cd.active)    #Currently Active sheet
active_sheet = cd['Logging']
print(cd.active.title)

active_sheet['A1'].value = 'Name'
active_sheet['B1'].value = 'Age'
active_sheet['A2'].value = 'Roger'
active_sheet['B2'].value = 30
active_sheet['B2'].fill = PatternFill('solid',fgColor='33FF5B')
active_sheet['A3'].value = 'Jack'
active_sheet['B3'].value = 40
active_sheet['B3'].fill = PatternFill('solid',fgColor='FF3364')

cd.save("Automation.xlsx")